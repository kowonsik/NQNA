#-*- coding: utf-8 -*-

import networkx as nx
import matplotlib.pyplot as plt
import numpy as np

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

from na_config import * 
import na_tools as nt
import na_build


import matplotlib

NewsQuoObjs=nt.loadObjectBinaryFast(NEWS_QUO_OBJ)

# graph is a list of edged in graph.  
#e.g. (23, 44), (39, 44), (39, 33), (44, 23), (44, 39), (44, 14), (14, 44), (33, 39), (39, 21), (39, 11), (44, 66), (44, 88), (33, 10)]
graph = []

inv_label = {}
labels={}
path_degree = {}
path_all = []

def draw_graph(graph, q_id, q_exemplar, graph_opt, argv_print):
    # extract nodes from graph
    #nodes = set([n1 for n1, n2 in graph] + [n2 for n1, n2 in graph])
    # Get list of nodes in edges
    nodes =list(set(zip(*graph)[0]+zip(*graph)[1]))
    # create networkx graph
    G=nx.Graph()

    for (qid, obj_) in NewsQuoObjs:
        # add nodes
        for node in nodes:
            if qid==node:
                labels[node]=str(node) +'\n'+ obj_.quotation[0:20]
                G.add_node(node)   # add nodes
    # add edges
    for edge in graph:
        G.add_edge(edge[0], edge[1])

    # draw graph
    #pos = nx.shell_layout(G)
    pos = nx.spring_layout(G)
    #pos = nx.random_layout(G)
    
    degree = nx.degree(G)

    nx.draw(G, pos, node_size=[v*100 for v in degree.values()], font_size=5, node_color='r')
    #nx.draw(G, pos, node_size=300, font_size=5, node_color='r', nodelist=q_exemplar_node)

    matplotlib.rc('font', family='AppleGothic')  # 폰트 설정
    
    if graph_opt == 'all':
        nx.draw_networkx_labels(G, pos, labels)
        plt.savefig(SAVE_ALL_CONNECTION) # save as png 
    else :
        nx.draw_networkx_labels(G, pos, labels)
        plt.savefig(SAVE_CONNECTION + graph_opt + '.png')  # save as png

    #plt.text(0,0,s='some text', bbox=dict(facecolor='red', alpha=0.5),horizontalalignment='center')graph = []

    # get all path
    if graph_opt == 'all':
        for k_s, v_s in q_id.iteritems():
            path_all=[]
            for k_t, v_t in q_id.iteritems():
                #if k_s == k_t:
                #    path_all.append('0')
                #else:
                for path in nx.all_simple_paths(G, source=v_s, target=v_t):
                    len_path = len(path)
                    exemplar_count = 0
                    for k_e, v_e in q_exemplar.iteritems():
                        for i in range(0, len_path):
                            if v_e == path[i]:
                                exemplar_count = exemplar_count+1
                    print v_s, v_t, path, q_exemplar, exemplar_count-1 
                    path_all.append([v_t,exemplar_count-1])
            path_degree[k_s] = path_all
	    #print k_s, path_degree[k_s]
    else:
        pass

    # show graph
    plt.show()

# same label finding
def find_same_label(q_label, argv_print):
    for k, v in q_label.iteritems():
        inv_label[v] = inv_label.get(v, [])
        inv_label[v].append(k)
    if argv_print == True:
        print inv_label
        print "========================"

def create_connect(q_id, q_exemplar, G_q, graph_opt='all', argv_print= True):
    # show all nodes wihtout label name
    if graph_opt == 'all':
        for i in range(0, G_q.shape[0]):
            for j in range(0, G_q.shape[1]):
                if i!=j:
                    if (G_q[i, j])==1:
                        graph.append((q_exemplar[i],q_exemplar[j]))    
        if argv_print == True:
            print "q_exemplar connected"
            print graph 
            print "========================"

        for k in range(0, len(inv_label)):
            v_size = len(inv_label[k])

            for h in range(0, v_size):
                if q_exemplar[k]!=q_id[inv_label[k][h]]:
                    #print str(q_exemplar[k]), str(q_id[inv_label[k][h]])
                    graph.append((q_exemplar[k],q_id[inv_label[k][h]]))
        if argv_print == True:
            print "q_exemplar & node connected"
            print graph 
            print "========================"

        draw_graph(graph, q_id, q_exemplar, graph_opt, argv_print)
    
    # in this case, graph_opt is 0 or 1 or 2, and so on, 
    # then show cluster-wise connect graph. 
    else :
        group_src =  inv_label[int(graph_opt)]
        #print group_src
        for k in range(0, len(group_src)):
            if q_exemplar[int(graph_opt)]!=q_id[group_src[k]]:
                #print str(q_exemplar[int(graph_opt)]), str(q_id[group_src[k]])
                graph.append((q_exemplar[int(graph_opt)],q_id[group_src[k]]))

        draw_graph(graph, q_id, q_exemplar, graph_opt, argv_print)


def save_result_excel(wb, RESULT_EXCEL, graph_opt):
    if graph_opt == 'all':
        wb.save(RESULT_ALL_EXCEL)
    else :
        wb.save(RESULT_EXCEL + graph_opt + '.xlsx')


def make_sheet1(wb,first,q_id,q_label,q_exemplar, RESULT_EXCEL, graph_opt, argv_print):
    # first sheet
    # write column name
    first.cell(row=1, column=1).value = 'id'
    first.cell(row=1, column=2).value = 'label'
    first.cell(row=1, column=3).value = 'exemplar'

    # ID input
    for k, v in q_id.iteritems():
        first.cell(row=k+2, column=1).value = v # start of 'k' is '2' 
        first.cell(row=k+2, column=3).value = 0 # init value = 0 

    # Level input
    for k, v in q_label.iteritems():
        first.cell(row=k+2, column=2).value = v # start of 'k' is '2' 


    # Exemplar input
    for k, v in q_id.iteritems():
        for j, h in q_exemplar.iteritems():
            if first.cell(row=k+1, column=1).value == h : 
                first.cell(row=k+1, column=3).value = 1 

    # save result excel
    save_result_excel(wb, RESULT_EXCEL, graph_opt)


def make_sheet2(wb, second, G_q, RESULT_EXCEL, graph_opt, argv_print):
    # second sheet
    # matrix input
    for i in range(0, G_q.shape[0]): 
        second.cell(row=1, column=i+2).value = i 

    for j in range(0, G_q.shape[0]): 
        second.cell(row=j+2, column=1).value = j 

    for m in range(0, G_q.shape[0]):
        for n in range(0, G_q.shape[1]):
            second.cell(row=m+2, column=n+2).value = G_q[m,n]

    # save result excel
    save_result_excel(wb, RESULT_EXCEL, graph_opt)


def excel_qid_insert(sheet, q_id, q_label):
	for k, v in q_id.iteritems():
		sheet.cell(row=k+2, column=1).value = v #  
		sheet.cell(row=1, column=k+2).value = v # 

        	# 0 depth path insert (same group)
        	for i, j in q_label.iteritems():
            		if q_label[k]==q_label[i]:
                		sheet.cell(row=k+2, column=i+2).value = 1 
			else :
                		sheet.cell(row=k+2, column=i+2).value = 0 

def make_sheet3(wb, third, q_id, q_label, G_q, RESULT_EXCEL, graph_opt, argv_print):

	for k, v in q_id.iteritems():
		for i in range(0, len(path_degree[k])):
			if path_degree[k][i][1] <= 1:
				#print v, i, path_degree[k][i][0], path_degree[k][i][1]
    				for m, g in q_id.iteritems():
					cellValue_id = third.cell(row=1, column=m+2).value
					if cellValue_id == path_degree[k][i][0]:
              					third.cell(row=k+2, column=m+2).value = 1 

    # save result excel
	save_result_excel(wb, RESULT_EXCEL, graph_opt)

def make_sheet4(wb, fourth, q_id, q_label, G_q, RESULT_EXCEL, graph_opt, argv_print):

	for k, v in q_id.iteritems():
		for i in range(0, len(path_degree[k])):
			if path_degree[k][i][1] <= 2:
				#print v, i, path_degree[k][i][0], path_degree[k][i][1]
    				for m, g in q_id.iteritems():
					cellValue_id = fourth.cell(row=1, column=m+2).value
					if cellValue_id == path_degree[k][i][0]:
              					fourth.cell(row=k+2, column=m+2).value = 1 

    # save result excel
	save_result_excel(wb, RESULT_EXCEL, graph_opt)

def make_sheet5(wb, fifth, q_id, q_label, G_q, RESULT_EXCEL, graph_opt, argv_print):

	for k, v in q_id.iteritems():
		for i in range(0, len(path_degree[k])):
			if path_degree[k][i][1] <= 3:
				#print v, i, path_degree[k][i][0], path_degree[k][i][1]
    				for m, g in q_id.iteritems():
					cellValue_id = fifth.cell(row=1, column=m+2).value
					if cellValue_id == path_degree[k][i][0]:
              					fifth.cell(row=k+2, column=m+2).value = 1 

    # save result excel
	save_result_excel(wb, RESULT_EXCEL, graph_opt)

def write_excel_result(q_id, q_label, q_exemplar, G_q, RESULT_EXCEL, SHEET_COUNT, graph_opt, argv_print):
    #print " creating excel result....."
    wb=Workbook()

    for i in range(0,SHEET_COUNT):
        ws = wb.create_sheet()

    sheetList = wb.get_sheet_names()
    first = wb.get_sheet_by_name(sheetList[0])
    second = wb.get_sheet_by_name(sheetList[1])
    third = wb.get_sheet_by_name(sheetList[2])
    fourth = wb.get_sheet_by_name(sheetList[3])
    fifth = wb.get_sheet_by_name(sheetList[4])

    excel_qid_insert(third, q_id, q_label)  # q_id insert and same q_id process in sheet 3 
    excel_qid_insert(fourth, q_id, q_label)  # q_id insert and same q_id process in sheet 3 
    excel_qid_insert(fifth, q_id, q_label)  # q_id insert and same q_id process in sheet 3 

    make_sheet1(wb, first, q_id, q_label, q_exemplar, RESULT_EXCEL, graph_opt, argv_print)
    make_sheet2(wb, second, G_q, RESULT_EXCEL, graph_opt, argv_print)
    make_sheet3(wb, third, q_id, q_label, G_q, RESULT_EXCEL, graph_opt, argv_print)
    make_sheet4(wb, fourth, q_id, q_label, G_q, RESULT_EXCEL, graph_opt, argv_print)
    make_sheet5(wb, fifth, q_id, q_label, G_q, RESULT_EXCEL, graph_opt, argv_print)


    

