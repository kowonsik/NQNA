# -*- coding: utf-8 -*-
"""
Created on Sun Sep 13 22:31:03 2015

@author: deokwooj
"""

"""
* Description 
- This contain all functions to extract data from excel files, and constracut
  data strcuture and store them as binary bin files for analysis inputs
"""
#########################################################
# News Article Excel file source 
#########################################################
#1. reference.xlsx ('분단'에 대한 자료 엑셀 파일)
#2. wholetable.xlsx (정보원 자료 엑셀 파일)
#3. table_define.xlsx : 정보원 정의
#########################################################
"""
        table_define.xlsx : 정보원 정의
        | infoSrc_ID                   | 정보원 ID |
        | name                         | 이름 |
        | orgName                      | 소속이름 |
        | type                         | 정보원 구분 |
        | position                     | 직함 |
        | etc_position                 | 기타 직함 정보 |
        | yearOfBirth                  | 생년 |
        | person_id(FK)                | 사전의 인물 ID |
        | code                         | 인물의 소속 분류 |
        | is_classified_paper_category | 신문 지면 정보에 의해 정보원의 분류되었는지 여부 |
        | INFOSRC_GLOBAL_ID            | 전기간
        | infosrc_id_whole    | 5 |
        | infosrc_id_day      | 2003/10/10_408 |
        | infosrc_name        | 김수행 |
        | infosrc_org         | 서울대 |
        | infosrc_type        | I |
        | infosrc_pos         | 교수 |
        | infosrc_code        | 13 |
        | infosrc_isclassified| \N |        에 걸친 UniqueID |
        
Each column is extracted from wholetable.xlsx (정보원 자료 엑셀 파일)
 """
 
""" 
type
| S | 익명 - 소속 없는 사람 |
| R | 익명 - 소속 있는 사람 |
| I | 실명 개인 - 소속 있음 |
| N | 무속속 실명 |
| O | 조직 |
| s | 성만 나와 있는 익명 |        
"""
# Deokwoo Jung 's update 23 Aug by jdw-2.  필수 입니다
# modules to be imported
# Loding common modules
from na_config import *
# Loading na_build specific modules
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook
from numpy import inf

######################################
# table_define.xlsx
# id : 정보원 ID
# name : 정보원 이름
# org : 정보원 조직
# srctype : 정보원 구분 {S,R,I,N,O,s}, (e.g. S ~ 익명 - 소속 없는 사람)
# pos : 정보원 직위
# code : 정보원 소속 분류, {헌법재판소, 재판부: 111 }, {검찰 : 211} 
# classified : 신문 지면 정보에 의해 정보원이 분류되어있는지 여부
######################################
class NewsSource:
    def __init__(self):
        self.id = None # uuid 
        self.date=dt.datetime(1999,12,31) #  quotations data,1999년 12월 31일 23시.
        self.name = None # name_seloadObjectBinaryFastt
        self.org = None # org_set
        self.srctype=None #~ {S,R,I,N,O,s}, (e.g. S ~ 익명 - 소속 없는 사람)
        self.pos = None #  Position  status_out!=True
        self.code=None # organization code
        self.classified=None # isclassified 
    def whoami(self): # print the current information for news source object
        for key in self.__dict__.items():
            print key[0],': ', key[1]\
            
######################################
# reference xlsx file
# 1 Reference sheet
#   - INFOSRC_NAME : 정보원 이름976.911 kB
#   - STN_CONTENT : 인용문이 들어간 문장
#   - ART_ID : 기사 ID
# 2 extraction sheet : 인용문 분리 후 명사 분리하고 정리
#   - 이름
#   - 인용문
#   - 명사
#   - 기사 ID
######################################
# Art.ID (meta_data_id) :"01101001[-->매체정보].20130527[-->날짜]100000112[-->기사ID] 
class NewsQuotation:
    def __init__(self):
        self.quotation_key =None # 4 digit number 
        self.article_id =None # 9 digit number 
        self.media_id = None  # 8 digit number string 
        self.date=dt.datetime(1999,12,31) #  quotations data,1999년 12월 31일 23시.
        self.news_src=NewsSource() # create NewsSource object
        self.quotation =None  # position, need utcto be initionalized by kkd_functions. 
        self.nounvec =None # position, need utcto be initionalized by kkd_functions. 
    def whoami(self):
        for key in self.__dict__.items():
            if key[0]=='news_src':
                print key[0],'name : ', key[1].name
            else:
                print key[0],': ', key[1]


# Building News Source Objects from WHOLETABLE_EXCEL and stores them into bin files.
def build_NewsSrcObjs(max_num_rows=inf):
    #TODO: define better status_flag 
    # ststus flag for output.     
    status_flag=True
    if max_num_rows <inf:
        print 'Building NewS Source Object up to ', max_num_rows, ' rows from ' +WHOLETABLE_EXCEL
    else:
        print 'Building NewS Source Object from ' +  WHOLETABLE_EXCEL

    try:
        wb=load_workbook(WHOLETABLE_EXCEL,use_iterators = True, read_only = True, keep_vba = False)
        #wb=load_workbook(WHOLETABLE_EXCEL,read_only = True, keep_vba = False)
        ws = wb.get_sheet_by_name(WHOLETABLE_SHEET)
    except:
        status_flag='Error in loading WHOLETABLE_EXCEL'
        return status_flag
    NewsSrcObjs=[]
    src_fld_name=\
    ['src_id','src_date','src_name','src_org','src_type','src_pos','src_code','src_clfd']
    for k,row in enumerate(ws.iter_rows(row_offset=1)):
        src_temp=NewsSource()
        if max_num_rows<k:
             break
        for (fld_, row_) in zip(src_fld_name,row):
            val_temp=row_.value
            if val_temp=='null':
                val_temp=None
            elif val_temp=='\N':
                val_temp='N'
            if val_temp!=None:
                if fld_=='src_id':
                    src_temp.id =val_temp
                elif fld_=='src_date':
                    val_temp=str(val_temp)[:10]
                    src_temp.date =dt.datetime(int(val_temp[:4]),int(val_temp[5:7]),int(val_temp[8:10]))
                elif fld_=='src_name':
                    src_temp.name =val_temp
                elif fld_=='src_org':
                    src_temp.org =val_temp
                elif fld_=='src_type':
                    src_temp.srctype =val_temp
                elif fld_=='src_pos':
                    src_temp.pos =val_temp
                elif fld_=='src_code':
                    src_temp.code =val_temp
                elif fld_=='src_clfd':                            
                    src_temp.classified =val_temp
                else:
                    warnings.warn("fld name not found")
                    status_flag='fld name not found'
                    return status_flag
        if src_temp.name!=None:
            NewsSrcObjs.append((k,src_temp))
    try: 
        nt.saveObjectBinaryFast(NewsSrcObjs, NEWS_SRC_OBJ)       
    except: 
        status_flag='Error in storing bin file'
        print status_flag
    return status_flag
    
def build_NewsQuoObjs(max_num_rows=inf):    
    # Load bin files for news sources and quotations 
    status_flag=True
    if max_num_rows <inf:
        print 'Building NewS Quotation Object up to ', max_num_rows, ' rows from ' +DICT_NEWS_INFO
    else:
        print 'Building NewS Quotation Object from' +DICT_NEWS_INFO 

    print 'Loading DICT_NEWS_INFO ...'
    # dict_news_info contains all quotations
    try:
        dict_news_info=nt.loadObjectBinaryFast(DICT_NEWS_INFO)
    except:
        status_flag='Error in loading '+DICT_NEWS_INFO
        print status_flag
        return status_flag

    NewsQuoObjs=[]
    ########################################################
    # Print all news dictionary information    
    ########################################################
    fld_name=('Name','Quatation', 'Nouns','Code')
    for k,(key, val) in enumerate(dict_news_info.iteritems()):
        quo_temp=NewsQuotation()
        quo_temp.quotation_key=str(key)
        if max_num_rows<k:
            break
        for i,(fld_, val_) in enumerate(zip(fld_name,list(val)[0])):
            print fld_
            if fld_=='Name':
                quo_temp.news_src.name=val_
            elif fld_=='Quatation':
                val_temp=val_
                quo_temp.quotation=val_
            elif fld_=='Nouns':
                quo_temp.nounvec=val_
            elif fld_=='Code':
                val_temp=str(val_)[1:]
                quo_temp.media_id=val_temp[:8]
                quo_temp.date=\
                dt.datetime(int(val_temp[9:13]),int(val_temp[13:15]),int(val_temp[15:17]),23)
                quo_temp.article_id=val_temp[17:]
            else:
                warnings.warn("fld name not found")
        NewsQuoObjs.append((k,quo_temp))
    try: 
        nt.saveObjectBinaryFast(NewsQuoObjs, NEWS_QUO_OBJ)   
    except: 
        status_flag='Error in storing bin file'
        print status_flag
    return status_flag


def na_build_main(SRC_OBJ=True,QUO_OBJ=True, argv_print=False):
    print 'running na_build.py....'
    print 'start to build news source object...'
    NewsSrcObjs=NewsQuoObjs=[]
    
    if SRC_OBJ==True:
        print 'start to build NewsSrcObjs....'
        try:
            print 'loading ' +NEWS_SRC_OBJ
            NewsSrcObjs=nt.loadObjectBinaryFast(NEWS_SRC_OBJ)
        except:
            print 'cannot find ' + NEWS_SRC_OBJ
            print 'buid it... '
            status_out=build_NewsSrcObjs()
            if status_out!=True:
                print status_out
            print 'job done, and stored news source objects...'
            NewsSrcObjs=nt.loadObjectBinaryFast(NEWS_SRC_OBJ)
        
        if argv_print==True:
            print '********************************************************************'
            print 'Print News Source Objects- NewsSrcObjs '
            for (sid, obj_) in NewsSrcObjs:
                print '==================================================='
                print 'News Source  ID : ', sid 
                print '==================================================='
                obj_.whoami()
            print '********************************************************************'
    else:
        print 'skip to build NewsSrcObjs....'

    if QUO_OBJ==True:
        print 'start to build NewsQuoObjs....'
        try:
            print 'loading ' +NEWS_QUO_OBJ
            NewsQuoObjs=nt.loadObjectBinaryFast(NEWS_QUO_OBJ)
        except:
            status_out=build_NewsQuoObjs()
            print status_out
            if status_out!=True:
                print status_outNEWS_SRC_OBJ
            print 'job done, and stored news quotation objects...'
            NewsQuoObjs=nt.loadObjectBinaryFast(NEWS_QUO_OBJ)
    
        if argv_print==True:
            print '********************************************************************'
            print 'Print News Quotation Objects- NewsQuoObjs '    
            for (qid, obj_) in NewsQuoObjs:
                print '==================================================='
                print 'Quotation ID : ', qid 
                print '==================================================='
                obj_.whoami()
            print '********************************************************************'
    else:
        print 'skip to build NewsQuoObjs....'
    
    return NewsSrcObjs, NewsQuoObjs

if __name__ == "__main__":
    na_build_main(argv_print=True)
