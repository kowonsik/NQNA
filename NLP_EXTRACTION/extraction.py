#-*- coding:utf-8 -*-
from konlpy.tag import Kkma
#from konlpy.utils import pprint
 
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter

from openpyxl import load_workbook

from na_tools import *
from na_config import *

def excel_noun():

	def excel_write(row_val, column_val, data):
		new_sheet.cell(row = row_val, column = column_val, value="%s" %data)

	wb=load_workbook(REFERENCE_EXCEL)

	sheetList = wb.get_sheet_names()
	sheet = wb.get_sheet_by_name(sheetList[0])
	row_count = sheet.get_highest_row()
	
	new_sheet = wb.create_sheet(title='extraction')
	
	news_info = {}
	
	for i in range(1, row_count):
		noun_val = ""
		full_qua = ""

		cellValue_name = sheet.cell(row=i, column=1).value
		cellValue = sheet.cell(row=i, column=2).value
		cellValue_id = sheet.cell(row=i, column=3).value

		# u201c 'LEFT DOUBLE QUOTATION MARK'
		# u201d 'RIGHT DOUBLE QUOTATION MARK'

		try :
			QUA = cellValue.count(u'\u201c')  # u201c 'LEFT DOUBLE QUOTATION MARK'
		except :
			continue 

		if QUA != -1:
			if QUA == 1 :
				START_QUA = cellValue.find(u"\u201c") + 1 # position of first quatation mark
				CELL_VALUE_LEN = len(cellValue)

				cellValue_re = cellValue[START_QUA:CELL_VALUE_LEN]
				END_QUA = cellValue_re.find(u"\u201d") # position of last quatation mark

				cellValue_final = cellValue_re[0:END_QUA]
				#print str(i) + "  "+ cellValue_name + "  "  + cellValue_final

				kkma = Kkma()
				#pprint (kkma.nouns(cellValue_final))
				s = (kkma.nouns(cellValue_final))

				for j in range(0,len(s)):
					noun_val = noun_val + s[j].encode('utf-8') + ','

				news_tuple=(cellValue_name, cellValue, noun_val, cellValue_id)
				news_info[i]={news_tuple}

				MyPrettyPrinter().pprint(news_info[i])

				excel_write(i, 1, cellValue_name)
				excel_write(i, 2, cellValue_final)
				excel_write(i, 3, noun_val)
				excel_write(i, 4, cellValue_id)

			elif QUA == 0 :
				#print str(i) + " " + cellValue
				ANOTHER_QUA = cellValue.find("\"") + 1 # position of first quatation mark
				ANOTHER_QUA_LEN = len(cellValue)

				another_cellValue = cellValue[ANOTHER_QUA:ANOTHER_QUA_LEN]
				ANOTHER_END_QUA = another_cellValue.find("\"")

				another_cellValue_final = another_cellValue[0:ANOTHER_END_QUA]
				#print str(i) + "  " + cellValue_name + "  " + another_cellValue_final
				kkma = Kkma()
				#pprint (kkma.nouns(cellValue_final))
				s = (kkma.nouns(another_cellValue_final))

				for j in range(0,len(s)):
					noun_val = noun_val + s[j].encode('utf-8') + ','

				news_tuple=(cellValue_name, cellValue, noun_val, cellValue_id)
				news_info[i]={news_tuple}

				MyPrettyPrinter().pprint(news_info[i])

				excel_write(i, 1, cellValue_name)
				excel_write(i, 2, another_cellValue_final)
				excel_write(i, 3, noun_val)
				excel_write(i, 4, cellValue_id)

			elif QUA > 1 :
				#print str(i) + " " + str(QUA)
				for q in range(0,QUA):
					arr = cellValue.split(u"\u201d")

					if arr is not None:
						try :
							arr_start_qua = arr[q].find(u"\u201c") + 1
						except :
							continue

						arr_len = len(arr[q]) 

						arr_cellValue = arr[q][arr_start_qua:arr_len]
						full_qua = full_qua + arr_cellValue

						kkma = Kkma()
						#pprint (kkma.nouns(cellValue_final))
						s = (kkma.nouns(arr_cellValue))

						for j in range(0,len(s)):
							noun_val = noun_val + s[j].encode('utf-8') + ','
							#print str(i) + " " + arr_cellValue

						news_tuple=(cellValue_name, cellValue, noun_val, cellValue_id)
						news_info[i]={news_tuple}

						MyPrettyPrinter().pprint(news_info[i])

						excel_write(i, 1, cellValue_name)
						excel_write(i, 2, full_qua)
						excel_write(i, 3, noun_val)
						excel_write(i, 4, cellValue_id)

	wb.save(REFERENCE_EXCEL)
	nt.saveObjectBinaryFast(news_info, DICT_NEWS_INFO) 


if __name__=="__main__":
	excel_noun()


